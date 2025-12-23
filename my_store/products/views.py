from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook
from django.db import transaction
from .models import Product, Category
from .serializers import ProductSerializer, ProductListSerializer, CategorySerializer


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return ProductListSerializer
        return ProductSerializer

    def get_queryset(self):
        queryset = Product.objects.all()
        category = self.request.query_params.get('category', None)
        search = self.request.query_params.get('search', None)
        is_active = self.request.query_params.get('is_active', None)

        if category:
            queryset = queryset.filter(category_id=category)
        if search:
            queryset = queryset.filter(name__icontains=search)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def update_stock(self, request, pk=None):
        product = self.get_object()
        quantity = request.data.get('quantity', 0)
        action_type = request.data.get('action', 'set')  # 'set', 'add', 'subtract'

        if action_type == 'set':
            product.stock = quantity
        elif action_type == 'add':
            product.stock += quantity
        elif action_type == 'subtract':
            product.stock = max(0, product.stock - quantity)

        product.save()
        serializer = self.get_serializer(product)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='import-excel', parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """
        Importe des produits depuis un fichier Excel.
        Le fichier doit contenir les colonnes 'designation' et 'quantite'.
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Aucun fichier fourni. Veuillez envoyer un fichier Excel.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        excel_file = request.FILES['file']
        
        # Vérifier que c'est un fichier Excel
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Le fichier doit être au format Excel (.xlsx ou .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Charger le workbook
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            
            # Trouver les colonnes 'designation' et 'quantite'
            header_row = None
            designation_col = None
            quantite_col = None
            
            # Chercher la ligne d'en-tête (premières 10 lignes)
            for row_idx in range(1, min(11, sheet.max_row + 1)):
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        cell_str = str(cell_value).strip().lower()
                        if 'designation' in cell_str or 'désignation' in cell_str:
                            designation_col = col_idx
                            header_row = row_idx
                        if 'quantite' in cell_str or 'quantité' in cell_str or 'qte' in cell_str:
                            quantite_col = col_idx
                            header_row = row_idx
            
            if not designation_col or not quantite_col:
                return Response(
                    {'error': 'Colonnes "designation" et "quantite" non trouvées dans le fichier Excel'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Traiter les lignes de données
            created_count = 0
            updated_count = 0
            errors = []
            
            start_row = header_row + 1 if header_row else 2
            
            with transaction.atomic():
                for row_idx in range(start_row, sheet.max_row + 1):
                    designation_cell = sheet.cell(row=row_idx, column=designation_col).value
                    quantite_cell = sheet.cell(row=row_idx, column=quantite_col).value
                    
                    # Ignorer les lignes vides
                    if not designation_cell:
                        continue
                    
                    designation = str(designation_cell).strip()
                    if not designation:
                        continue
                    
                    # Convertir la quantité en entier
                    try:
                        if quantite_cell is None:
                            quantite = 0
                        else:
                            quantite = int(float(str(quantite_cell)))
                    except (ValueError, TypeError):
                        errors.append(f"Ligne {row_idx}: Quantité invalide '{quantite_cell}'")
                        continue
                    
                    # Chercher si le produit existe déjà (par nom)
                    product, created = Product.objects.get_or_create(
                        name=designation,
                        defaults={
                            'stock': quantite,
                            'price': 0.00,  # Prix par défaut
                            'is_active': True,
                            'created_by': request.user
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        # Mettre à jour le stock
                        product.stock = quantite
                        product.save()
                        updated_count += 1

            return Response({
                'message': f'Import terminé avec succès',
                'created': created_count,
                'updated': updated_count,
                'errors': errors if errors else None
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'Erreur lors du traitement du fichier Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

