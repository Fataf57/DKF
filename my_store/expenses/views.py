from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Expense, ExpenseCategory
from .serializers import ExpenseSerializer, ExpenseListSerializer, ExpenseCategorySerializer


class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer
    permission_classes = [IsAuthenticated]


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return ExpenseListSerializer
        return ExpenseSerializer

    def get_queryset(self):
        queryset = Expense.objects.all()
        category = self.request.query_params.get('category', None)
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)
        search = self.request.query_params.get('search', None)

        if category:
            queryset = queryset.filter(category_id=category)
        if date_from:
            queryset = queryset.filter(expense_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(expense_date__lte=date_to)
        if search:
            queryset = queryset.filter(description__icontains=search)

        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def export_report(self, request):
        """Génère un rapport Excel des dépenses pour une période donnée"""
        date_from = request.query_params.get('date_from', None)
        date_to = request.query_params.get('date_to', None)

        queryset = Expense.objects.all()
        if date_from:
            queryset = queryset.filter(expense_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(expense_date__lte=date_to)

        queryset = queryset.order_by('expense_date')

        # Créer un workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport des Dépenses"

        # En-têtes
        headers = ['Date', 'Description', 'Catégorie', 'Montant', 'Méthode de Paiement', 'Notes']
        ws.append(headers)

        # Style des en-têtes
        header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Données
        total_general = 0
        for expense in queryset:
            row = [
                expense.expense_date.strftime('%d/%m/%Y'),
                expense.description,
                expense.category.name if expense.category else '-',
                float(expense.amount),
                expense.get_payment_method_display(),
                expense.notes or '-'
            ]
            ws.append(row)
            total_general += float(expense.amount)

        # Ajouter le total
        ws.append([])
        ws.append(['TOTAL GÉNÉRAL', '', '', total_general, '', ''])

        # Style du total
        total_row = ws[ws.max_row]
        for cell in total_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Ajuster la largeur des colonnes
        column_widths = [12, 40, 20, 12, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Créer la réponse HTTP
        filename = f"rapport_depenses_{date_from or 'all'}_{date_to or 'all'}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

