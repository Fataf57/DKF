from rest_framework import viewsets, status, generics
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Sum, Count
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Sale, SaleItem, OutOfStockSale
from .serializers import (
    SaleSerializer, SaleCreateSerializer, SaleListSerializer, 
    SaleUpdateSerializer, SaleItemSerializer, OutOfStockSaleSerializer
)


class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return SaleCreateSerializer
        elif self.action == 'list':
            return SaleListSerializer
        elif self.action in ['update', 'partial_update']:
            return SaleUpdateSerializer
        return SaleSerializer

    def get_queryset(self):
        from django.utils import timezone
        from datetime import datetime
        from django.db.models import Q
        
        # Ne pas utiliser prefetch_related pour éviter les erreurs de conversion Decimal
        queryset = Sale.objects.select_related('customer')
        customer = self.request.query_params.get('customer', None)
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)

        if customer:
            queryset = queryset.filter(customer_id=customer)
        if date_from:
            # Utiliser __date pour comparer uniquement la date (ignorer l'heure)
            # Cela inclura toutes les ventes à partir de cette date, même s'il n'y a pas de vente à cette date exacte
            queryset = queryset.filter(sale_date__date__gte=date_from)
        if date_to:
            # Utiliser __date pour comparer uniquement la date (ignorer l'heure)
            queryset = queryset.filter(sale_date__date__lte=date_to)

        return queryset
    
    def list(self, request, *args, **kwargs):
        """Surcharge pour gérer les erreurs de conversion Decimal en utilisant values()"""
        import logging
        from rest_framework.response import Response
        from rest_framework import status
        from decimal import Decimal, InvalidOperation
        
        logger = logging.getLogger(__name__)
        
        try:
            queryset = self.filter_queryset(self.get_queryset())
            
            # Récupérer uniquement les IDs pour éviter toute conversion Decimal lors de values()
            sale_ids = list(queryset.values_list('id', flat=True))
            
            sales_data = []
            for sale_id in sale_ids:
                try:
                    # Récupérer la vente complète en utilisant only() pour éviter de charger total_amount immédiatement
                    # On le chargera manuellement avec gestion d'erreur
                    sale = Sale.objects.only(
                        'id', 'customer_id', 'sale_date', 'payment_method', 
                        'notes', 'created_at', 'customer'
                    ).get(id=sale_id)
                    
                    # Charger total_amount séparément avec gestion d'erreur
                    # Utiliser une requête SQL brute pour éviter la conversion Decimal
                    try:
                        from django.db import connection
                        with connection.cursor() as cursor:
                            cursor.execute("SELECT total_amount FROM sales_sale WHERE id = %s", [sale_id])
                            row = cursor.fetchone()
                            if row and row[0] is not None:
                                try:
                                    # Essayer de convertir en Decimal
                                    total_amount_str = str(Decimal(str(row[0])).quantize(Decimal('0.01')))
                                except (ValueError, InvalidOperation, TypeError):
                                    total_amount_str = '0.00'
                            else:
                                total_amount_str = '0.00'
                    except Exception as e:
                        logger.error(f"Erreur lors de la récupération de total_amount pour la vente {sale_id}: {e}")
                        total_amount_str = '0.00'
                    
                    # Charger les items sans unit_price pour éviter la conversion Decimal
                    # On récupérera unit_price directement depuis l'objet
                    items_ids = sale.items.values_list('id', flat=True)
                    items_data = []
                    
                    for item_id in items_ids:
                        try:
                            # Récupérer l'item complet
                            from .models import SaleItem
                            item = SaleItem.objects.filter(id=item_id).first()
                            if not item:
                                continue
                            
                            # Récupérer le produit
                            from products.models import Product
                            product = item.product
                            
                            # Gérer unit_price directement depuis l'objet avec gestion d'erreur
                            try:
                                unit_price_attr = item.unit_price
                                if unit_price_attr is None:
                                    unit_price = Decimal('0.00')
                                elif isinstance(unit_price_attr, Decimal):
                                    unit_price = unit_price_attr
                                else:
                                    unit_price = Decimal(str(unit_price_attr))
                            except (ValueError, InvalidOperation, TypeError, AttributeError):
                                unit_price = Decimal('0.00')
                            
                            quantity = item.quantity if item.quantity is not None else 0
                            try:
                                quantity = int(quantity)
                            except (ValueError, TypeError):
                                quantity = 0
                            
                            subtotal = Decimal(str(quantity)) * unit_price
                            
                            items_data.append({
                                'id': item.id,
                                'product': product.id if product else None,
                                'product_name': product.name if product else '-',
                                'quantity': quantity,
                                'unit_price': str(unit_price.quantize(Decimal('0.01'))),
                                'subtotal': str(subtotal.quantize(Decimal('0.01')))
                            })
                        except Exception as item_error:
                            logger.error(f"Erreur avec l'item {item_id}: {item_error}", exc_info=True)
                            continue
                    
                    # Construire les données de la vente
                    customer = sale.customer
                    # total_amount_str a déjà été récupéré plus haut avec gestion d'erreur
                    
                    sale_data = {
                        'id': sale.id,
                        'customer': sale.customer_id,
                        'customer_name': customer.full_name if customer else None,
                        'sale_date': sale.sale_date.isoformat() if sale.sale_date else None,
                        'total_amount': total_amount_str,
                        'payment_method': sale.payment_method,
                        'payment_method_display': sale.get_payment_method_display() if hasattr(sale, 'get_payment_method_display') else sale.payment_method,
                        'notes': sale.notes or '',
                        'items': items_data,
                        'items_count': len(items_data),
                        'created_at': sale.created_at.isoformat() if sale.created_at else None,
                    }
                    sales_data.append(sale_data)
                    
                except Exception as sale_error:
                    logger.error(f"Erreur avec la vente {sale_id}: {sale_error}", exc_info=True)
                    continue
            
            return Response(sales_data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des ventes: {e}", exc_info=True)
            return Response(
                {'error': 'Erreur lors de la récupération des ventes. Vérifiez les logs pour plus de détails.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_destroy(self, instance):
        """Restaure les stocks avant de supprimer la vente"""
        from django.db import transaction
        
        with transaction.atomic():
            # Restaurer les stocks pour chaque item
            for item in instance.items.all():
                product = item.product
                product.stock += item.quantity
                product.save(update_fields=['stock'])
            
            # Supprimer les enregistrements de ventes hors stock associés
            OutOfStockSale.objects.filter(sale=instance).delete()
            
            # Supprimer la vente (les items seront supprimés en cascade)
            instance.delete()
    

    @action(detail=False, methods=['get'])
    def export_report(self, request):
        """Génère un rapport Excel des ventes pour une période donnée"""
        date_from = request.query_params.get('date_from', None)
        date_to = request.query_params.get('date_to', None)

        queryset = Sale.objects.all()
        if date_from:
            queryset = queryset.filter(sale_date__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(sale_date__date__lte=date_to)

        queryset = queryset.order_by('sale_date')

        # Créer un workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport des Ventes"

        # En-têtes
        headers = ['Date', 'Heure', 'ID Vente', 'Produit', 'Quantité', 'Prix Unitaire', 'Total', 'Méthode de Paiement']
        ws.append(headers)

        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Données
        total_general = 0
        for sale in queryset:
            sale_date = sale.sale_date.date()
            sale_time = sale.sale_date.time()
            for item in sale.items.all():
                row = [
                    sale_date.strftime('%d/%m/%Y'),
                    sale_time.strftime('%H:%M'),
                    sale.id,
                    item.product.name,
                    item.quantity,
                    float(item.unit_price),
                    float(item.subtotal),
                    sale.get_payment_method_display()
                ]
                ws.append(row)
                total_general += float(item.subtotal)

        # Ajouter le total
        ws.append([])
        ws.append(['TOTAL GÉNÉRAL', '', '', '', '', '', total_general, ''])

        # Style du total
        total_row = ws[ws.max_row]
        for cell in total_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Ajuster la largeur des colonnes
        column_widths = [12, 10, 10, 30, 10, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Créer la réponse HTTP
        filename = f"rapport_ventes_{date_from or 'all'}_{date_to or 'all'}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'])
    def check_stock(self, request):
        """Vérifie les stocks avant l'enregistrement d'une vente"""
        items = request.data.get('items', [])
        stock_issues = []
        
        for item in items:
            product_id = item.get('product')
            quantity = item.get('quantity', 0)
            
            if not product_id or not quantity:
                continue
                
            from products.models import Product
            try:
                product = Product.objects.get(id=product_id)
                if product.stock < quantity:
                    stock_issues.append({
                        'product_id': product_id,
                        'product_name': product.name,
                        'requested': quantity,
                        'available': product.stock,
                        'shortage': quantity - product.stock
                    })
            except Product.DoesNotExist:
                stock_issues.append({
                    'product_id': product_id,
                    'product_name': 'Produit introuvable',
                    'requested': quantity,
                    'available': 0,
                    'shortage': quantity
                })
        
        return Response({
            'has_issues': len(stock_issues) > 0,
            'issues': stock_issues
        })


class OutOfStockSaleListView(generics.ListAPIView):
    """Liste des ventes hors stock"""
    queryset = OutOfStockSale.objects.all()
    serializer_class = OutOfStockSaleSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = OutOfStockSale.objects.all().order_by('-created_at')
        return queryset

